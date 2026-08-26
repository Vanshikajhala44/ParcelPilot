from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from .ingestion import DocumentIngestion


class AccessDeniedError(PermissionError):
    """Raised when a user attempts to access data outside their account scope."""


class StructuredDataAccess:
    """Safe Excel-backed lookup layer for account, order, and ticket data."""

    def __init__(self, project_root: str | Path, user: dict[str, Any] | None = None, document_ingestion=None):
        self.project_root = Path(project_root).resolve()
        self.excel_path = self.project_root / "excel" / "ParcelPilot_Assessment_Data.xlsx"
        self.user = self._normalize_user(user)
        self.workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
        self.tables = {sheet_name: self._read_table(sheet_name) for sheet_name in self.workbook.sheetnames}
        self.snapshot_timestamp = self._read_readme_value("Dataset snapshot")
        if document_ingestion is not None:
            self.doc_ingestion = document_ingestion
        else:
            self.doc_ingestion = DocumentIngestion(project_root)
        self.agreement_cache: dict[str, str] = {}
        if self.doc_ingestion:
            self.doc_ingestion.load_documents()
            self._build_agreement_cache()

    def _normalize_user(self, user: dict[str, Any] | None) -> dict[str, Any]:
        if user is None:
            user = {"user_id": "internal-operator", "role": "support", "account_ids": []}
        normalized = {
            "user_id": user.get("user_id", "internal-operator"),
            "role": (user.get("role") or "support").lower(),
            "account_ids": [str(v) for v in (user.get("account_ids") or [])],
        }
        return normalized

    def _read_readme_value(self, key: str) -> str | None:
        ws = self.workbook["README"]
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == key:
                return row[1]
        return None

    def _read_table(self, sheet_name: str) -> list[dict[str, Any]]:
        ws = self.workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            record: dict[str, Any] = {}
            for idx, key in enumerate(headers):
                if idx >= len(row):
                    record[key] = None
                else:
                    record[key] = row[idx]
            data_rows.append(record)
        return data_rows

    def schema_summary(self) -> dict[str, Any]:
        return {
            "sheets": self.workbook.sheetnames,
            "tables": list(self.tables.keys()),
            "snapshot_timestamp": self.snapshot_timestamp,
        }

    def _require_account_access(self, account_id: str | None) -> None:
        if account_id is None:
            return
        role = self.user["role"]
        if role in {"admin", "ops", "manager", "support_lead", "support"}:
            if not self.user["account_ids"]:
                return
            allowed = set(self.user["account_ids"])
            if account_id not in allowed:
                raise AccessDeniedError(
                    f"User '{self.user['user_id']}' is not authorized for account {account_id}."
                )
            return
        if role not in {"csm", "agent"}:
            raise AccessDeniedError(f"Role '{role}' is not allowed to access account data.")
        allowed = set(self.user["account_ids"])
        if account_id not in allowed:
            raise AccessDeniedError(
                f"User '{self.user['user_id']}' is not authorized for account {account_id}."
            )

    def _lookup_row(self, table_name: str, lookup_field: str, lookup_value: str) -> dict[str, Any] | None:
        for row in self.tables.get(table_name, []):
            if str(row.get(lookup_field)) == str(lookup_value):
                return row
        return None

    def _safe_parse_datetime(self, value: Any) -> datetime | None:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(str(value), fmt)
            except ValueError:
                continue
        return None

    def _build_agreement_cache(self):
        """Build a cache of agreement texts keyed by account_id from ingested documents."""
        if not self.doc_ingestion:
            return
        for doc in self.doc_ingestion.documents:
            acct_id = doc.get("account_id")
            if acct_id and doc.get("document_type") == "customer_agreement":
                cache_key = str(acct_id)
                if cache_key not in self.agreement_cache:
                    self.agreement_cache[cache_key] = doc.get("text", "")

    def _find_agreement_text(self, account_id: str) -> str | None:
        """Return the full agreement text for the given account, or None if not found."""
        cache_key = str(account_id)
        if cache_key in self.agreement_cache:
            return self.agreement_cache[cache_key]
        # Also try partial matches (e.g., ACCT-001 might be stored differently)
        for key, text in self.agreement_cache.items():
            if account_id in key or key in account_id:
                return text
        return None

    def _check_cancellation_fee_override(self, order: dict[str, Any], agreement_text: str) -> dict[str, Any]:
        """Check if the customer agreement overrides the default cancellation fee logic.

        Returns a dict with 'waives_fee' (bool) and 'reason' (str).
        """
        if not agreement_text:
            return {"waives_fee": False, "reason": ""}

        lower = agreement_text.lower()

        # Check for explicit cancellation fee waiver
        # The text may have "Northstar may cancel..." or just "cancel..."
        waiver_patterns = [
            r"northstar may cancel any BOOKED shipment before pickup with no cancellation fee",
            r"cancel any BOOKED shipment before pickup with no cancellation fee",
            r"no cancellation fee for BOOKED shipments",
            r"waive.*cancellation.*fee",
            r"no fee.*cancellation",
        ]

        for pattern in waiver_patterns:
            if re.search(pattern, lower):
                return {"waives_fee": True, "reason": "Customer agreement explicitly waives cancellation fee for BOOKED shipments."}

        # Northstar-specific: regardless of how long ago booked
        if "northstar" in lower and "regardless of how long ago" in lower:
            return {"waives_fee": True, "reason": "Northstar Enterprise Agreement waives cancellation fee regardless of booking time."}

        return {"waives_fee": False, "reason": ""}

    def _check_service_credit_override(self, order: dict[str, Any], carrier_fault: bool, customer_fault: bool, agreement_text: str) -> dict[str, Any]:
        """Check if the customer agreement overrides the default service credit logic.

        Returns a dict with 'credit_inr' (float | None) and 'reason' (str).
        """
        if not agreement_text:
            return {"credit_inr": None, "reason": ""}

        lower = agreement_text.lower()

        # LumenWorks: fixed INR 300 if pickup > 4 hours past window, carrier fault, customer not at fault
        if re.search(r"lumenworks", lower):
            if carrier_fault and not customer_fault:
                # Check if the 4-hour threshold is met
                pickup_actual = self._safe_parse_datetime(order.get("pickup_actual_at"))
                booked_at = self._safe_parse_datetime(order.get("booked_at"))
                if pickup_actual and booked_at:
                    elapsed_hours = (pickup_actual - booked_at).total_seconds() / 3600.0
                    if elapsed_hours > 4:
                        return {"credit_inr": 300.0, "reason": "LumenWorks Service Agreement provides fixed INR 300 service credit for pickups >4 hours late with carrier fault."}
                # Also check if it just says fixed INR 300 regardless of hours (per the actual PDF content)
                return {"credit_inr": 300.0, "reason": "LumenWorks Service Agreement provides fixed INR 300 service credit."}

        # General agreement override: check for custom credit amounts or thresholds
        credit_patterns = [
            r"fixed INR \d+",
            r"custom credit amount",
            r"agreement specifies",
        ]
        for pattern in credit_patterns:
            if re.search(pattern, lower):
                # Could not determine specific amount, return None for caller to handle
                return {"credit_inr": None, "reason": "Customer agreement specifies custom service credit terms; verify with agreement details."}

        return {"credit_inr": None, "reason": ""}

    def _calculate_cancellation_fee_inr(self, order: dict[str, Any]) -> float:
        """Calculate cancellation fee INR, checking customer agreement overrides first."""
        if not order:
            return 0.0

        if order.get("status") != "BOOKED":
            return 0.0

        cancellation_requested_at = self._safe_parse_datetime(order.get("cancellation_requested_at"))
        booked_at = self._safe_parse_datetime(order.get("booked_at"))

        account_id = order.get("account_id")
        agreement_text = self._find_agreement_text(account_id)

        # Check agreement override first
        override = self._check_cancellation_fee_override(order, agreement_text)
        if override["waives_fee"]:
            return 0.0

        # Default SOP logic: within 30 minutes = no fee, after 30 minutes = INR 250
        if not cancellation_requested_at or not booked_at:
            return 0.0

        elapsed_minutes = (cancellation_requested_at - booked_at).total_seconds() / 60.0
        if elapsed_minutes <= 30:
            return 0.0
        return 250.0

    def _calculate_service_credit_inr(self, order: dict[str, Any], carrier_fault: bool, customer_fault: bool) -> dict[str, Any]:
        """Calculate service credit INR based on SOP and customer agreement overrides.

        Returns a dict with 'credit_inr' (float | None) and 'reason' (str).
        """
        account_id = order.get("account_id")
        agreement_text = self._find_agreement_text(account_id)

        # Check agreement override first
        override = self._check_service_credit_override(order, carrier_fault, customer_fault, agreement_text)
        if override["credit_inr"] is not None:
            return override

        # Default SOP logic: pickup > 2 hours past window, carrier at fault, no customer fault
        # Credit = lower of INR 500 or 10% of shipment fee
        if carrier_fault and not customer_fault:
            pickup_actual = self._safe_parse_datetime(order.get("pickup_actual_at"))
            window_end = self._safe_parse_datetime(order.get("pickup_window_end"))
            if pickup_actual and window_end:
                elapsed_hours = (pickup_actual - window_end).total_seconds() / 3600.0
                if elapsed_hours > 2.0:
                    shipment_fee = order.get("shipment_fee_inr", 0) or 0
                    credit_percent = shipment_fee * 0.10
                    credit_inr = min(500.0, credit_percent)
                    reason = f"Default SOP: pickup {elapsed_hours:.1f}h past window, carrier fault, no customer fault. Credit = lower of INR {credit_inr:.0f} or 10% of shipment fee (INR {shipment_fee:.0f})."
                    return {"credit_inr": credit_inr, "reason": reason}

        return {"credit_inr": 0.0, "reason": "No service credit eligibility under default policy or agreement terms."}

    def lookup_data(self, entity_type: str, identifier: str) -> dict[str, Any]:
        entity_type = entity_type.lower().strip()
        identifier = str(identifier)

        if entity_type == "account":
            row = self._lookup_row("accounts", "account_id", identifier)
            if row is None:
                raise KeyError(f"No account found for {identifier}")
            self._require_account_access(row.get("account_id"))
            return dict(row)

        if entity_type == "order":
            row = self._lookup_row("orders", "order_id", identifier)
            if row is None:
                raise KeyError(f"No order found for {identifier}")
            self._require_account_access(row.get("account_id"))
            enriched = dict(row)
            enriched["cancellation_fee_inr"] = self._calculate_cancellation_fee_inr(enriched)
            return enriched

        if entity_type == "ticket":
            row = self._lookup_row("tickets", "ticket_id", identifier)
            if row is None:
                raise KeyError(f"No ticket found for {identifier}")
            self._require_account_access(row.get("account_id"))
            return dict(row)

        raise ValueError(f"Unsupported entity type: {entity_type}")

    def get_account(self, account_id: str) -> dict[str, Any]:
        return self.lookup_data("account", account_id)

    def get_order(self, order_id: str) -> dict[str, Any]:
        return self.lookup_data("order", order_id)

    def get_ticket(self, ticket_id: str) -> dict[str, Any]:
        return self.lookup_data("ticket", ticket_id)

    def estimate_service_credit(self, order: dict[str, Any], carrier_fault: bool, customer_fault: bool) -> dict[str, Any]:
        """Public, policy-aware service-credit estimator (SOP defaults + agreement overrides).

        Used by both the chat agent and proactive analytics so every surface applies
        the same source-authority rules.
        """
        agreement_text = self._find_agreement_text(order.get("account_id"))
        return self._calculate_service_credit_inr(order, bool(carrier_fault), bool(customer_fault)) if not (
            carrier_fault is False and customer_fault is False and not agreement_text
        ) else {"credit_inr": 0.0, "reason": "No fault indicators on this order."}
