import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\vansh\OneDrive\Desktop\project_new\excel\ParcelPilot_Assessment_Data.xlsx', read_only=True, data_only=True)
print('Sheet names:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
        print(f'\nSheet: {sheet_name}')
        print(f'Headers ({len(headers)}): {headers}')
        print(f'Data rows: {len(rows)-1}')
        for row in rows[1:3]:
            print(f'  {row}')